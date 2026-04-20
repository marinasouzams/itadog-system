#!/usr/bin/env python3
"""
Extract all ITADOG data into structured JSON for SQL generation.
"""
import os
import re
import glob
import json
from openpyxl import load_workbook

DIRS = [
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/arquivositadog (1)",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/rearquivositadog (1)",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/rearquivositadog (2)",
]

def cv(ws, r, c):
    try:
        v = ws.cell(row=r, column=c).value
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip().lstrip("'")
            return v if v else None
        return v
    except:
        return None

def is_order_sheet(name):
    name = str(name).strip()
    return bool(re.match(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$', name))

def parse_date(name):
    name = name.strip().replace('/', '.').replace('-', '.')
    parts = name.split('.')
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 2:
            y = '20' + y
        try:
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        except:
            pass
    return name

def extract_client(ws):
    info = {"name": None, "cnpj": None, "city": None, "phone": None, "email": None}
    for r in range(1, 20):
        for c in range(1, 15):
            raw = ws.cell(row=r, column=c).value
            if not raw or not isinstance(raw, str):
                continue
            v = raw.strip().lstrip("'")
            if not v:
                continue
            vl = v.lower()
            nv_raw = ws.cell(row=r, column=c+1).value
            nv = str(nv_raw).strip().lstrip("'") if nv_raw else None

            if 'nome:' in vl:
                m = re.search(r'nome[:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['name'] = m.group(1).strip()
            elif 'cnpj' in vl:
                m = re.search(r'cnpj[.:\s]*(.+)', v, re.IGNORECASE)
                val = m.group(1).strip() if m else None
                if val and val not in ['', '.']:
                    info['cnpj'] = val
                elif nv:
                    info['cnpj'] = nv
            elif 'cidade' in vl:
                m = re.search(r'cidade[:\s]+(.+)', v, re.IGNORECASE)
                val = m.group(1).strip() if m else None
                if val:
                    info['city'] = val
                elif nv:
                    info['city'] = nv
            elif any(x in vl for x in ['tel:', 'fone:', 'celular:', 'whats:']):
                m = re.search(r'(?:tel|fone|celular|whats)[.:\s]+(.+)', v, re.IGNORECASE)
                val = m.group(1).strip() if m else None
                if val:
                    info['phone'] = val
                elif nv:
                    info['phone'] = nv
            elif 'email' in vl or 'e-mail' in vl:
                m = re.search(r'e?-?mail[:\s]+(.+)', v, re.IGNORECASE)
                val = m.group(1).strip() if m else None
                if val:
                    info['email'] = val
                elif nv:
                    info['email'] = nv
    return info

def extract_products(ws):
    products = []
    seen = set()
    for r in range(5, ws.max_row + 1):
        code = cv(ws, r, 1)
        desc = cv(ws, r, 2)
        if not code and not desc:
            continue
        if isinstance(code, str):
            if any(x in code.lower() for x in ['código', 'total', 'chegue', 'prazo', 'desconto', 'nome']):
                continue
        if isinstance(desc, str):
            if any(x in desc.lower() for x in ['descrição', 'total', 'produto']):
                continue
        cost = cv(ws, r, 8)
        sale = cv(ws, r, 9)
        if cost is None and sale is None:
            continue
        key = (str(code), str(desc))
        if key in seen:
            continue
        seen.add(key)
        products.append({
            "code": str(code).strip() if code else None,
            "description": str(desc).strip() if desc else None,
            "cost_price": round(float(cost), 2) if isinstance(cost, (int, float)) else None,
            "sale_price": round(float(sale), 2) if isinstance(sale, (int, float)) else None,
        })
    return products

def extract_order(ws, sheet_name):
    date = parse_date(sheet_name)
    items = []
    discount = None
    payment_terms = None
    payment_dates = []

    for r in range(1, ws.max_row + 1):
        # Scan all cols for payment/discount text
        for c in range(1, ws.max_column + 1):
            val = cv(ws, r, c)
            if val and isinstance(val, str):
                vl = val.lower().strip()
                if re.search(r'chegue\s*\d', vl) or re.search(r'^\d+[-/]\d+[-/]\d+', vl):
                    payment_terms = val.strip()
                elif re.match(r'^\d+[-]\d+', vl) and 'dias' not in vl:
                    payment_terms = val.strip()
                elif 'à vista' in vl or 'avista' in vl or '30 dias' in vl or 'prazo' in vl:
                    payment_terms = val.strip()
                if 'desconto' in vl:
                    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', val)
                    if m:
                        discount = float(m.group(1).replace(',', '.'))
                # Payment date patterns
                if re.match(r'^\d{2}\.\d{2}\.\d{2,4}$', vl):
                    payment_dates.append(val.strip())

        code = cv(ws, r, 1)
        desc = cv(ws, r, 2)
        sale = cv(ws, r, 9)
        qty = cv(ws, r, 10)
        total = cv(ws, r, 11)

        if code is None and qty is None:
            continue
        if isinstance(code, str):
            if any(x in code.lower() for x in ['código', 'total', 'chegue', 'prazo', 'desconto', 'nome', 'itadog']):
                continue
        if isinstance(qty, (int, float)) and qty > 0 and code is not None:
            items.append({
                "code": str(code).strip() if code else None,
                "description": str(desc).strip() if desc else None,
                "unit_price": round(float(sale), 2) if isinstance(sale, (int, float)) else None,
                "qty": int(qty),
                "total": round(float(total), 2) if isinstance(total, (int, float)) else None,
            })

    return {
        "sheet_name": sheet_name,
        "date": date,
        "discount_pct": discount,
        "payment_terms": payment_terms,
        "payment_dates": list(dict.fromkeys(payment_dates)),  # unique, preserve order
        "items": items,
    }

def process_file(filepath):
    filename = os.path.basename(filepath)
    client_name_from_file = re.sub(r'^PEDIDO\s+AGROPECUARIA\s+', '', filename, flags=re.IGNORECASE)
    client_name_from_file = re.sub(r'^PEDIDO\s+', '', client_name_from_file, flags=re.IGNORECASE)
    client_name_from_file = re.sub(r'\.xlsx$', '', client_name_from_file, flags=re.IGNORECASE).strip()

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return {"filename": filename, "error": str(e)}

    client_info = None
    products = []
    orders = []
    other_sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sname = sheet_name.strip()

        if sname.upper() in ['PEDIDO MODELO', 'MODELO', 'TABELA DE PRECOS', 'TABELA']:
            client_info = extract_client(ws)
            products = extract_products(ws)

        elif is_order_sheet(sname):
            order = extract_order(ws, sname)
            if order['items']:  # only include orders with items
                orders.append(order)

        elif sname.lower() not in ['plan2', 'plan3', 'sheet1', 'sheet2', 'sheet3', 'planilha1', 'planilha2']:
            # Non-standard sheet - check if it has order-like data
            # Try to extract as old-format order
            rows_data = []
            for r in range(1, ws.max_row + 1):
                row = {}
                for c in range(1, ws.max_column + 1):
                    v = cv(ws, r, c)
                    if v is not None:
                        row[c] = v
                if row:
                    rows_data.append({"row": r, "data": row})
            if rows_data:
                other_sheets.append({"sheet": sname, "rows": rows_data})

    if not client_info:
        client_info = {}
    if not client_info.get('name'):
        client_info['name'] = client_name_from_file

    return {
        "filename": filename,
        "client": client_info,
        "products": products,
        "orders": orders,
        "other_sheets": other_sheets,
    }

# Collect unique files
all_files = {}
for d in DIRS:
    for f in glob.glob(os.path.join(d, "*.xlsx")):
        bname = os.path.basename(f)
        if bname not in all_files:
            all_files[bname] = f

results = []
for bname in sorted(all_files.keys()):
    result = process_file(all_files[bname])
    results.append(result)

# Save to JSON
out_path = "/Users/marinajuliadesouza/Documents/sitema-itadog/extracted_data.json"
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)

print(f"Saved {len(results)} files to {out_path}")
print(f"Total file size: {os.path.getsize(out_path)} bytes")

# Print summary
for r in results:
    if "error" in r:
        print(f"  ERROR {r['filename']}: {r['error']}")
    else:
        n_orders = len(r.get('orders', []))
        n_products = len(r.get('products', []))
        n_other = len(r.get('other_sheets', []))
        client = r.get('client', {})
        print(f"  {r['filename']}")
        print(f"    Client: {client.get('name')} | CNPJ: {client.get('cnpj')} | City: {client.get('city')} | Phone: {client.get('phone')}")
        print(f"    Products: {n_products} | Orders: {n_orders} | Other sheets: {n_other}")
        for o in r.get('orders', []):
            items_count = len(o['items'])
            print(f"    Order {o['sheet_name']} ({o['date']}): {items_count} items | terms={o['payment_terms']} | discount={o['discount_pct']}% | pay_dates={o['payment_dates']}")
        if n_other > 0:
            for os_ in r.get('other_sheets', []):
                print(f"    Other sheet '{os_['sheet']}': {len(os_['rows'])} rows")
