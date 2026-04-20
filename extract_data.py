#!/usr/bin/env python3
"""
Extract all data from ITADOG Excel files for SQL INSERT generation.
"""
import os
import re
import json
from openpyxl import load_workbook
from datetime import datetime

FILES = [
    "ALINE LIMA",
    "CARLOS AUGUSTO",
    "CRISTINA",
    "DANIELA MARCELINO",
    "FERNANDA GOULART",
    "GEOVANA",
    "GIULIA MAIA",
    "INGRID",
    "JESSICA OLIVEIRA",
    "LETICIA GUIMARAES",
    "LIVIA FREITAS",
    "MARISA",
    "NATHALIA AMARAL",
    "NATHALIA PONTES",
    "NELI",
    "SILVIA",
    "THALIA",
    "VIVIANE PIACINI",
]

BASE_DIR = "/Users/marinajuliadesouza/Documents/sitema-itadog/planilhas"

def cell_val(ws, row, col):
    """Get cell value safely."""
    try:
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        if isinstance(v, str):
            return v.strip()
        return v
    except:
        return None

def find_label(ws, label, max_rows=30, max_cols=15):
    """Find a cell containing label text, return (row, col)."""
    label_lower = label.lower()
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if label_lower in cell.value.lower():
                    return (cell.row, cell.column)
    return (None, None)

def extract_header(ws):
    """Extract client info from header rows (first ~15 rows)."""
    info = {
        "name": None, "cnpj": None, "city": None,
        "phone": None, "email": None, "address": None
    }

    # Dump first 20 rows to understand structure
    header_data = []
    for r in range(1, 25):
        row_vals = []
        for c in range(1, 15):
            v = cell_val(ws, r, c)
            if v is not None:
                row_vals.append((c, v))
        if row_vals:
            header_data.append((r, row_vals))

    return header_data

def extract_client_info(ws):
    """Try to extract structured client info from common patterns."""
    info = {}

    # Scan first 20 rows for labels
    for r in range(1, 25):
        for c in range(1, 15):
            v = cell_val(ws, r, c)
            if v and isinstance(v, str):
                vl = v.lower().strip()
                # Check for label patterns and grab value from adjacent cell
                next_val = cell_val(ws, r, c + 1)
                if next_val is None:
                    next_val = cell_val(ws, r + 1, c)

                if 'cliente' in vl or 'client' in vl or 'nome' in vl:
                    if next_val:
                        info['name'] = str(next_val)
                elif 'cnpj' in vl or 'cpf' in vl:
                    if next_val:
                        info['cnpj'] = str(next_val)
                elif 'cidade' in vl or 'city' in vl or 'munic' in vl:
                    if next_val:
                        info['city'] = str(next_val)
                elif 'tel' in vl or 'fone' in vl or 'cel' in vl or 'whats' in vl:
                    if next_val:
                        info['phone'] = str(next_val)
                elif 'email' in vl or 'e-mail' in vl:
                    if next_val:
                        info['email'] = str(next_val)
                elif 'endere' in vl or 'rua' in vl or 'av.' in vl:
                    if next_val:
                        info['address'] = str(next_val)

    return info

def is_order_sheet(name):
    """Check if sheet name looks like a date (e.g., '10.02.26', '05/03/2026')."""
    name = str(name).strip()
    # Patterns: DD.MM.YY, DD.MM.YYYY, DD/MM/YY, DD/MM/YYYY, or date-like
    patterns = [
        r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$',
        r'^\d{2}\.\d{2}\.\d{2,4}$',
        r'^\d{2}/\d{2}/\d{2,4}$',
    ]
    for p in patterns:
        if re.match(p, name):
            return True
    return False

def parse_date_from_sheet(name):
    """Parse date from sheet name like '10.02.26' or '05/03/2026'."""
    name = name.strip().replace('/', '.').replace('-', '.')
    parts = name.split('.')
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 2:
            y = '20' + y
        try:
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        except:
            pass
    return name

def extract_products(ws):
    """Extract products from PEDIDO MODELO sheet."""
    products = []

    # Print all content to understand structure
    all_rows = []
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = cell_val(ws, r, c)
            if v is not None:
                row_vals.append((c, v))
        if row_vals:
            all_rows.append((r, row_vals))

    return all_rows

def extract_order(ws):
    """Extract order items from an order sheet."""
    all_rows = []
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = cell_val(ws, r, c)
            if v is not None:
                row_vals.append((c, v))
        if row_vals:
            all_rows.append((r, row_vals))
    return all_rows

def process_file(filename):
    filepath = os.path.join(BASE_DIR, filename + ".xlsx")
    print(f"\n{'='*80}")
    print(f"FILE: {filename}")
    print(f"{'='*80}")

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"ERROR loading {filepath}: {e}")
        return

    print(f"SHEETS: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sname = sheet_name.strip()

        print(f"\n--- SHEET: '{sname}' ---")

        if sname.upper() in ['PEDIDO MODELO', 'MODELO', 'TABELA', 'PRODUTOS', 'PRODUTO']:
            print("[PRODUCT/MODEL SHEET]")
            rows = extract_products(ws)
            for r, row_vals in rows:
                print(f"  Row {r}: {row_vals}")

        elif is_order_sheet(sname):
            date = parse_date_from_sheet(sname)
            print(f"[ORDER SHEET] date={date}")
            rows = extract_order(ws)
            for r, row_vals in rows:
                print(f"  Row {r}: {row_vals}")

        else:
            # Could be client info sheet or something else
            print(f"[OTHER SHEET]")
            rows = []
            for r in range(1, min(ws.max_row + 1, 50)):
                row_vals = []
                for c in range(1, ws.max_column + 1):
                    v = cell_val(ws, r, c)
                    if v is not None:
                        row_vals.append((c, v))
                if row_vals:
                    rows.append((r, row_vals))
                    print(f"  Row {r}: {row_vals}")

for fname in FILES:
    process_file(fname)
