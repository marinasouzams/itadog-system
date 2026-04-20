#!/usr/bin/env python3
"""
Generate complete SQL INSERT statements from all ITADOG Excel files.
"""
import os
import re
import glob
import json
from openpyxl import load_workbook

DIRS = [
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/arquivositadog (1)",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/rearquivositadog (1)",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/rearquivositadog (2)",
]

def cv(ws, r, c):
    try:
        v = ws.cell(row=r, column=c).value
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip().lstrip("'").strip()
            return v if v else None
        return v
    except:
        return None

def is_date_sheet(name):
    name = str(name).strip()
    return bool(re.match(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$', name))

def is_old_date_sheet(name):
    """Sheets like '29.10', '07.11' — DD.MM only"""
    name = str(name).strip()
    return bool(re.match(r'^\d{2}\.\d{2}$', name))

def parse_date(name, base_year='2025'):
    name = name.strip().replace('/', '.').replace('-', '.')
    parts = name.split('.')
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 2:
            y = '20' + y
        try:
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        except:
            pass
    elif len(parts) == 2:
        d, m = parts
        # Guess year from context - old sheets from 2025
        return f"{base_year}-{m.zfill(2)}-{d.zfill(2)}"
    return name

def q(v):
    """Quote for SQL"""
    if v is None:
        return 'NULL'
    s = str(v).replace("'", "''")
    return f"'{s}'"

def extract_client_info(ws):
    """Extract client info from any sheet."""
    info = {"name": None, "cnpj": None, "city": None, "phone": None, "email": None, "address": None, "cep": None}

    # From any sheet: extract patterns
    for r in range(1, 25):
        for c in range(1, 15):
            raw = ws.cell(row=r, column=c).value
            if not raw or not isinstance(raw, str):
                continue
            v = raw.strip().lstrip("'").strip()
            if not v:
                continue
            vl = v.lower()
            nv_raw = ws.cell(row=r, column=c+1).value
            nv = str(nv_raw).strip().lstrip("'").strip() if nv_raw and isinstance(nv_raw, str) else None

            if 'nome:' in vl:
                m = re.search(r'nome[:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['name'] = m.group(1).strip()
            if 'cnpj' in vl:
                # Try to get value from same cell first (embedded)
                m = re.search(r'cnpj[.:\s]+([0-9][0-9./\-,]+)', v, re.IGNORECASE)
                if m:
                    raw_cnpj = m.group(1).strip().replace(',', '.')
                    # Validate looks like a CNPJ (has enough digits)
                    digits = re.sub(r'[^0-9]', '', raw_cnpj)
                    if len(digits) >= 8:
                        info['cnpj'] = raw_cnpj
                elif nv:
                    # Value in next column
                    digits = re.sub(r'[^0-9]', '', nv)
                    if len(digits) >= 8:
                        info['cnpj'] = nv
            if 'cep' in vl:
                m = re.search(r'cep[.:\s]+([0-9\-]+)', v, re.IGNORECASE)
                if m:
                    info['cep'] = m.group(1).strip()
            if re.match(r'^tel[.:\s]', vl) or re.match(r'^telefax:', vl) or re.match(r'^celular:', vl) or re.match(r'^whats:', vl):
                # Only pick real phone numbers, not "TABELA DE PREÇOS ATUAL"
                m = re.search(r'(?:telefax|tel|celular|whats)[.:\s]+([0-9][0-9\s\-]+)', v, re.IGNORECASE)
                if m and m.group(1).strip() and len(re.sub(r'[^0-9]', '', m.group(1))) >= 6:
                    info['phone'] = m.group(1).strip()
            if 'email' in vl or 'e-mail' in vl:
                m = re.search(r'e?-?mail[:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['email'] = m.group(1).strip()
            if re.match(r'^(r\.|av\.|rua|avenida)', vl):
                if not info['address']:
                    info['address'] = v

    return info

def extract_products_new_format(ws):
    """Auto-detect format and extract products from PEDIDO MODELO sheet."""
    products = []
    seen = set()

    # Detect format from row 5
    r5c1 = cv(ws, 5, 1)
    r5c8 = cv(ws, 5, 8)
    is_new = r5c1 and isinstance(r5c1, str) and 'código' in r5c1.lower()
    is_old = (not is_new) and r5c8 and isinstance(r5c8, str) and 'preço' in r5c8.lower()

    for r in range(5, ws.max_row + 1):
        if is_new:
            code = cv(ws, r, 1)
            desc = cv(ws, r, 2)
            cost = cv(ws, r, 8)
            sale = cv(ws, r, 9)
        elif is_old:
            # Old format: desc col1, cost col7, sale col8
            code = None
            desc = cv(ws, r, 1)
            cost = cv(ws, r, 7)
            sale = cv(ws, r, 8)
        else:
            # Default to new format
            code = cv(ws, r, 1)
            desc = cv(ws, r, 2)
            cost = cv(ws, r, 8)
            sale = cv(ws, r, 9)

        if not code and not desc:
            continue
        if isinstance(code, str):
            code_l = code.lower()
            if any(x in code_l for x in ['código', 'total', 'chegue', 'prazo', 'desconto', 'nome', 'itadog', 'coluna']):
                continue
        if isinstance(desc, str):
            desc_l = desc.lower()
            if any(x in desc_l for x in ['descrição', 'total geral', 'produto', 'itadog']):
                continue
        if cost is None and sale is None:
            continue
        if not isinstance(cost, (int, float)) and not isinstance(sale, (int, float)):
            continue

        # For old format, skip header-like text in col1
        if is_old and isinstance(desc, str):
            desc_l = desc.lower()
            if any(x in desc_l for x in ['tel:', 'cnpj', 'nome:', 'preço', 'qta', 'vlr', 'tabela']):
                continue

        key = (str(code), str(desc))
        if key in seen:
            continue
        seen.add(key)

        code_str = str(code).strip() if code is not None else None
        # For old format where code is description
        if is_old and code_str is None:
            code_str = None
            desc_str = str(desc).strip() if desc is not None else None
        else:
            desc_str = str(desc).strip() if desc is not None else None

        products.append({
            "code": code_str,
            "description": desc_str,
            "cost_price": round(float(cost), 2) if isinstance(cost, (int, float)) else None,
            "sale_price": round(float(sale), 2) if isinstance(sale, (int, float)) else None,
        })
    return products

def extract_order_new_format(ws, sheet_name):
    """New format order sheet: code col1, desc col2, cost col8, sale col9, qty col10, total col11."""
    date = parse_date(sheet_name)
    items = []
    discount = None
    payment_terms = None
    payment_dates = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = cv(ws, r, c)
            if val and isinstance(val, str):
                vl = val.lower().strip()
                # Payment terms
                if re.search(r'chegue\s*\d', vl):
                    payment_terms = val.strip()
                elif re.match(r'^nf\s+prazo', vl):
                    payment_terms = val.strip()
                elif re.match(r'^prazo\s+\d', vl):
                    payment_terms = val.strip()
                elif re.match(r'^\d+[-]\d+', vl) and not re.match(r'^\d{2}\.\d{2}', vl):
                    payment_terms = val.strip()
                elif any(x in vl for x in ['à vista', 'avista', 'boleto', 'pix agendado']):
                    payment_terms = val.strip()
                elif re.match(r'^\d+ dias', vl):
                    payment_terms = val.strip()
                # Discount
                if 'desconto' in vl:
                    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', val)
                    if m:
                        discount = float(m.group(1).replace(',', '.'))
                # Payment dates
                if re.match(r'^\d{2}\.\d{2}\.\d{2,4}$', vl):
                    payment_dates.append(val.strip())

        code = cv(ws, r, 1)
        desc = cv(ws, r, 2)
        sale = cv(ws, r, 9)
        qty = cv(ws, r, 10)
        total = cv(ws, r, 11)

        if code is None or qty is None:
            continue
        if isinstance(code, str):
            if any(x in code.lower() for x in ['código', 'total', 'chegue', 'prazo', 'desconto', 'nome', 'itadog']):
                continue
        if isinstance(qty, (int, float)) and qty > 0:
            items.append({
                "code": str(code).strip() if code is not None else None,
                "description": str(desc).strip() if desc is not None else None,
                "unit_price": round(float(sale), 2) if isinstance(sale, (int, float)) else None,
                "qty": int(qty),
                "total": round(float(total), 2) if isinstance(total, (int, float)) else None,
            })

    return {
        "sheet_name": sheet_name,
        "date": date,
        "discount_pct": discount,
        "payment_terms": payment_terms,
        "payment_dates": list(dict.fromkeys(payment_dates)),
        "items": items,
        "format": "new",
    }

def extract_order_old_format(ws, sheet_name):
    """Old format: desc col1, cost col7 or col8, sale col8, qty col9, total col10.
    Date found in row 2 col 10."""
    # Get actual date from row 2
    actual_date_raw = None
    for r in range(1, 6):
        for c in range(8, 15):
            val = cv(ws, r, c)
            if val and isinstance(val, str):
                vl = val.strip()
                if re.match(r'^\d{2}\.\d{2}\.\d{2,4}$', vl) or re.match(r'^\d{2}\.\d{2}\.\d{2}$', vl):
                    actual_date_raw = vl
                    break

    if actual_date_raw:
        date = parse_date(actual_date_raw)
    else:
        date = parse_date(sheet_name, '2025')

    items = []
    discount = None
    payment_terms = None
    payment_dates = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = cv(ws, r, c)
            if val and isinstance(val, str):
                vl = val.lower().strip()
                if any(x in vl for x in ['à vista', 'avista', 'prazo', '30 dias']):
                    payment_terms = val.strip()
                if re.match(r'^\d{2}\.\d{2}\.\d{2,4}$', vl):
                    payment_dates.append(val.strip())

        desc = cv(ws, r, 1)
        # Old format: col8=sale, col9=qty, col10=total (some have cost in col7)
        sale8 = cv(ws, r, 8)
        sale9_maybe = cv(ws, r, 9)
        qty = cv(ws, r, 9)
        total = cv(ws, r, 10)

        # Detect columns: if col9 has 'PREÇO' label, it's the sale price col
        # Row 5 usually has headers: col8='PREÇO', col9='QTA', col10='VLR'
        # So sale=col8, qty=col9, total=col10

        if desc is None:
            continue
        if isinstance(desc, str):
            desc_l = desc.lower()
            if any(x in desc_l for x in ['itadog', 'nome:', 'cnpj', 'tel:', 'preço', 'qta', 'vlr', 'total', 'razão']):
                continue
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue
        if not isinstance(sale8, (int, float)):
            continue

        items.append({
            "code": None,
            "description": str(desc).strip() if desc else None,
            "unit_price": round(float(sale8), 2),
            "qty": int(qty),
            "total": round(float(total), 2) if isinstance(total, (int, float)) else None,
        })

    return {
        "sheet_name": sheet_name,
        "date": date,
        "discount_pct": discount,
        "payment_terms": payment_terms,
        "payment_dates": list(dict.fromkeys(payment_dates)),
        "items": items,
        "format": "old",
    }

def extract_cadastro(ws):
    """Extract full client info from CADASTRO sheet."""
    info = {"name": None, "razao_social": None, "address": None, "city": None, "cep": None, "cnpj": None, "phone": None, "ie": None}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(row=r, column=c).value
            if not raw or not isinstance(raw, str):
                continue
            v = raw.strip().lstrip("'").strip()
            if not v:
                continue
            vl = v.lower()

            if 'razão social' in vl or 'razao social' in vl:
                continue  # label
            elif 'destinário' in vl or 'destinario' in vl or 'remetente' in vl:
                continue  # label
            elif re.match(r'^cnpj[.:\s]', vl):
                m = re.search(r'cnpj[.:\s]+([0-9./\-]+)', v, re.IGNORECASE)
                if m:
                    info['cnpj'] = m.group(1).strip()
            elif re.match(r'^insc[.:\s]', vl):
                m = re.search(r'insc[.:\s]+est[.:\s]+([0-9]+)', v, re.IGNORECASE)
                if m:
                    info['ie'] = m.group(1).strip()
            elif re.match(r'^cep[.:\s]', vl):
                m = re.search(r'cep[.:\s]+([0-9\-]+)', v, re.IGNORECASE)
                if m:
                    info['cep'] = m.group(1).strip()
            elif re.match(r'^tel[.:\s]', vl):
                m = re.search(r'tel[.:\s]+([0-9\s\-]+)', v, re.IGNORECASE)
                if m:
                    info['phone'] = m.group(1).strip()
            elif re.match(r'^(r\.|av\.|rua|avenida)', vl):
                info['address'] = v
            elif any(x in vl for x in ['(sc)', '(pr)', '(sp)', '(rj)']):
                info['city'] = v
            elif not any(char.isdigit() for char in v) and len(v) > 3 and r > 1:
                # Likely a name
                if not info['name']:
                    info['name'] = v
                elif not info['razao_social']:
                    info['razao_social'] = v
    return info

def process_file(filepath):
    filename = os.path.basename(filepath)
    client_name_from_file = re.sub(r'^PEDIDO\s+AGROPECUARIA\s+', '', filename, flags=re.IGNORECASE)
    client_name_from_file = re.sub(r'^PEDIDO\s+', '', client_name_from_file, flags=re.IGNORECASE)
    client_name_from_file = re.sub(r'\.xlsx$', '', client_name_from_file, flags=re.IGNORECASE).strip()

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return None

    client_info = {}
    cadastro_info = {}
    products = []
    orders = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sname = sheet_name.strip()

        if sname.upper() in ['PEDIDO MODELO', 'MODELO']:
            ci = extract_client_info(ws)
            client_info.update({k: v for k, v in ci.items() if v is not None})
            prods = extract_products_new_format(ws)
            if prods:
                products = prods

        elif sname.upper() == 'CADASTRO':
            cadastro_info = extract_cadastro(ws)

        elif is_date_sheet(sname):
            # Auto-detect format by checking row 5
            r5c1 = cv(ws, 5, 1)
            r5c8 = cv(ws, 5, 8)
            if r5c1 and isinstance(r5c1, str) and 'código' in r5c1.lower():
                order = extract_order_new_format(ws, sname)
            elif r5c8 and isinstance(r5c8, str) and 'preço' in r5c8.lower():
                order = extract_order_old_format(ws, sname)
            else:
                # Try new format first
                order = extract_order_new_format(ws, sname)
                if not order['items']:
                    order = extract_order_old_format(ws, sname)
            if order['items']:
                orders.append(order)

        elif is_old_date_sheet(sname) or re.match(r'^\d{2}\.\d{2}$', sname) or re.match(r'^\d{2}\.\d{2}\s', sname):
            order = extract_order_old_format(ws, sname)
            if order['items']:
                orders.append(order)

        elif sname.upper() == 'ETIQUETA':
            ci = extract_cadastro(ws)
            for k, v in ci.items():
                if v and not cadastro_info.get(k):
                    cadastro_info[k] = v

        elif sname.lower() not in ['plan2', 'plan3', 'sheet1', 'sheet2', 'sheet3', 'planilha1', 'planilha2']:
            # Named sheets like "PEDIDO 10.09.2025", "Plan1" with order data
            # Auto-detect format
            r5c1 = cv(ws, 5, 1)
            r5c8 = cv(ws, 5, 8)
            if r5c8 and isinstance(r5c8, str) and 'preço' in r5c8.lower():
                order = extract_order_old_format(ws, sname)
            elif r5c1 and isinstance(r5c1, str) and 'código' in r5c1.lower():
                order = extract_order_new_format(ws, sname)
            else:
                order = extract_order_old_format(ws, sname)
            if order['items']:
                orders.append(order)

    # Merge client info: cadastro takes precedence for name/cnpj/city
    if cadastro_info.get('name') and not client_info.get('name'):
        client_info['name'] = cadastro_info['name']
    if cadastro_info.get('cnpj'):
        client_info['cnpj'] = cadastro_info['cnpj']
    if cadastro_info.get('city'):
        client_info['city'] = cadastro_info['city']
    if cadastro_info.get('phone'):
        client_info['phone'] = cadastro_info['phone']
    if cadastro_info.get('address'):
        client_info['address'] = cadastro_info['address']
    if cadastro_info.get('cep'):
        client_info['cep'] = cadastro_info['cep']
    if cadastro_info.get('ie'):
        client_info['ie'] = cadastro_info['ie']

    if not client_info.get('name'):
        client_info['name'] = client_name_from_file

    return {
        "filename": filename,
        "client_key": client_name_from_file,
        "client": client_info,
        "products": products,
        "orders": orders,
    }

# Collect unique files (main dir first to avoid duplicates)
all_files = {}
for d in DIRS:
    for f in glob.glob(os.path.join(d, "*.xlsx")):
        bname = os.path.basename(f)
        if bname not in all_files:
            all_files[bname] = f

results = []
for bname in sorted(all_files.keys()):
    result = process_file(all_files[bname])
    if result:
        results.append(result)

# -----------------------------------------------------------------------
# Build master product catalog (deduplicated across all files)
# -----------------------------------------------------------------------
all_products = {}  # code -> product
for r in results:
    for p in r.get('products', []):
        code = p['code']
        if code and code not in all_products:
            all_products[code] = p

# -----------------------------------------------------------------------
# Generate SQL
# -----------------------------------------------------------------------
sql_lines = []
sql_lines.append("-- ================================================================")
sql_lines.append("-- ITADOG: Generated SQL INSERT statements")
sql_lines.append("-- ================================================================")
sql_lines.append("")

# 1. CLIENTS
sql_lines.append("-- ----------------------------------------------------------------")
sql_lines.append("-- CLIENTS")
sql_lines.append("-- ----------------------------------------------------------------")
sql_lines.append("INSERT INTO clients (name, cnpj, city, phone, email, address, cep, state_registration, file_key) VALUES")

client_rows = []
for r in results:
    c = r['client']
    row = (
        f"  ({q(c.get('name'))}, {q(c.get('cnpj'))}, {q(c.get('city'))}, "
        f"{q(c.get('phone'))}, {q(c.get('email'))}, {q(c.get('address'))}, "
        f"{q(c.get('cep'))}, {q(c.get('ie'))}, {q(r['client_key'])})"
    )
    client_rows.append(row)
sql_lines.append(",\n".join(client_rows) + ";")
sql_lines.append("")

# 2. PRODUCTS
sql_lines.append("-- ----------------------------------------------------------------")
sql_lines.append("-- PRODUCTS (deduplicated catalog)")
sql_lines.append("-- ----------------------------------------------------------------")
sql_lines.append("INSERT INTO products (code, description, cost_price, sale_price) VALUES")
prod_rows = []
for code, p in sorted(all_products.items()):
    row = (
        f"  ({q(p['code'])}, {q(p['description'])}, "
        f"{p['cost_price'] if p['cost_price'] is not None else 'NULL'}, "
        f"{p['sale_price'] if p['sale_price'] is not None else 'NULL'})"
    )
    prod_rows.append(row)
sql_lines.append(",\n".join(prod_rows) + ";")
sql_lines.append("")

# 3. ORDERS + ORDER ITEMS
sql_lines.append("-- ----------------------------------------------------------------")
sql_lines.append("-- ORDERS")
sql_lines.append("-- ----------------------------------------------------------------")

order_id = 1
for r in results:
    client_key = r['client_key']
    for order in r.get('orders', []):
        discount = order['discount_pct'] if order['discount_pct'] is not None else 'NULL'
        payment_terms = q(order['payment_terms'])
        payment_dates = q(", ".join(order['payment_dates'])) if order['payment_dates'] else 'NULL'
        sql_lines.append(
            f"INSERT INTO orders (id, client_key, order_date, sheet_name, discount_pct, payment_terms, payment_dates, format) VALUES"
        )
        sql_lines.append(
            f"  ({order_id}, {q(client_key)}, {q(order['date'])}, {q(order['sheet_name'])}, "
            f"{discount}, {payment_terms}, {payment_dates}, {q(order.get('format', 'new'))});"
        )
        sql_lines.append("")

        if order['items']:
            sql_lines.append(f"INSERT INTO order_items (order_id, product_code, description, qty, unit_price, total) VALUES")
            item_rows = []
            for item in order['items']:
                item_rows.append(
                    f"  ({order_id}, {q(item['code'])}, {q(item['description'])}, "
                    f"{item['qty']}, "
                    f"{item['unit_price'] if item['unit_price'] is not None else 'NULL'}, "
                    f"{item['total'] if item['total'] is not None else 'NULL'})"
                )
            sql_lines.append(",\n".join(item_rows) + ";")
            sql_lines.append("")

        order_id += 1

# Save SQL
out_sql = "/Users/marinajuliadesouza/Documents/sitema-itadog/itadog_data.sql"
with open(out_sql, 'w', encoding='utf-8') as f:
    f.write("\n".join(sql_lines))

# Save JSON too
out_json = "/Users/marinajuliadesouza/Documents/sitema-itadog/extracted_data_full.json"
with open(out_json, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)

print(f"SQL saved to: {out_sql} ({os.path.getsize(out_sql):,} bytes)")
print(f"JSON saved to: {out_json} ({os.path.getsize(out_json):,} bytes)")
print(f"\nSUMMARY:")
print(f"  Clients: {len(results)}")
print(f"  Unique products in catalog: {len(all_products)}")
total_orders = sum(len(r['orders']) for r in results)
total_items = sum(len(o['items']) for r in results for o in r['orders'])
print(f"  Total orders: {total_orders}")
print(f"  Total order items: {total_items}")
print()

for r in results:
    orders = r.get('orders', [])
    c = r['client']
    print(f"\n{'='*60}")
    print(f"CLIENT: {c.get('name')}")
    print(f"  CNPJ: {c.get('cnpj')} | City: {c.get('city')} | Phone: {c.get('phone')}")
    print(f"  Address: {c.get('address')} | CEP: {c.get('cep')} | IE: {c.get('ie')}")
    print(f"  Products: {len(r.get('products', []))} | Orders: {len(orders)}")
    for o in orders:
        print(f"    Order '{o['sheet_name']}' ({o['date']}) [{o.get('format','?')}]: "
              f"{len(o['items'])} items | terms={o['payment_terms']} | discount={o['discount_pct']}%")
        for item in o['items']:
            print(f"      - {item['code']} | {item['description']} | qty={item['qty']} | price={item['unit_price']} | total={item['total']}")
