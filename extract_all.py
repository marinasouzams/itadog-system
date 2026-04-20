#!/usr/bin/env python3
"""
Extract all data from ITADOG Excel files for SQL INSERT generation.
Reads every file, every sheet, and extracts client info, products, and orders.
"""
import os
import re
import glob
from openpyxl import load_workbook

DIRS = [
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/arquivositadog (1)",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/rearquivositadog (1)",
    "/Users/marinajuliadesouza/Downloads/ARQUIVOS ITADOG/rearquivositadog (2)",
]

def cv(ws, r, c):
    """Get cell value, strip strings."""
    try:
        v = ws.cell(row=r, column=c).value
        if v is None:
            return None
        if isinstance(v, str):
            v = v.strip().lstrip("'")
        return v
    except:
        return None

def is_order_sheet(name):
    name = str(name).strip()
    return bool(re.match(r'^\d{1,2}[./]\d{1,2}[./]\d{2,4}$', name))

def parse_date(name):
    name = name.strip().replace('/', '.').replace('-', '.')
    parts = name.split('.')
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 2:
            y = '20' + y
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return name

def extract_client_from_sheet(ws):
    """Extract client name, CNPJ, city, phone, email from any sheet."""
    info = {"name": None, "cnpj": None, "city": None, "phone": None, "email": None}
    for r in range(1, 20):
        for c in range(1, 15):
            raw = ws.cell(row=r, column=c).value
            if not raw or not isinstance(raw, str):
                continue
            v = raw.strip().lstrip("'")
            vl = v.lower()
            # Try to get value from next column or same cell after ':'
            nv = cv(ws, r, c+1)

            if 'nome:' in vl or vl.startswith('nome '):
                # The name is often in the same cell after "NOME: "
                m = re.search(r'nome[:\s]+(.+)', v, re.IGNORECASE)
                if m:
                    info['name'] = m.group(1).strip()
                elif nv:
                    info['name'] = str(nv).strip().lstrip("'")
            elif 'cnpj' in vl:
                m = re.search(r'cnpj[.:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['cnpj'] = m.group(1).strip()
                elif nv:
                    info['cnpj'] = str(nv).strip().lstrip("'")
            elif 'cidade' in vl or 'munic' in vl:
                m = re.search(r'cidade[:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['city'] = m.group(1).strip()
                elif nv:
                    info['city'] = str(nv).strip().lstrip("'")
            elif any(x in vl for x in ['tel', 'fone', 'celular', 'whats']):
                m = re.search(r'(?:tel|fone|celular|whats)[.:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['phone'] = m.group(1).strip()
                elif nv:
                    info['phone'] = str(nv).strip().lstrip("'")
            elif 'email' in vl or 'e-mail' in vl:
                m = re.search(r'e?-?mail[:\s]+(.+)', v, re.IGNORECASE)
                if m and m.group(1).strip():
                    info['email'] = m.group(1).strip()
                elif nv:
                    info['email'] = str(nv).strip().lstrip("'")
    return info

def extract_products_from_sheet(ws):
    """Extract products: code, description, cost_price, sale_price."""
    products = []
    # Header is on row 5: CÓDIGO | DESCRIÇÃO | ... | PREÇO(cost col8) | PREÇO(sale col9) | QTDE | VLR
    # Products start at row 6
    for r in range(5, ws.max_row + 1):
        code = cv(ws, r, 1)
        desc = cv(ws, r, 2)
        if code is None and desc is None:
            continue
        if not code and not desc:
            continue
        # Skip header rows
        if isinstance(code, str) and 'código' in code.lower():
            continue
        if isinstance(desc, str) and ('descrição' in desc.lower() or 'total' in desc.lower()):
            continue
        # Get prices - col 8 = cost, col 9 = sale price
        cost = cv(ws, r, 8)
        sale = cv(ws, r, 9)
        # Skip rows without prices (likely totals or notes)
        if cost is None and sale is None:
            continue
        # Skip if code looks like a label
        if isinstance(code, str) and any(x in code.lower() for x in ['total', 'chegue', 'prazo', 'desconto']):
            continue
        products.append({
            "code": str(code).strip().lstrip("'") if code else None,
            "description": str(desc).strip().lstrip("'") if desc else None,
            "cost_price": float(cost) if cost and isinstance(cost, (int, float)) else None,
            "sale_price": float(sale) if sale and isinstance(sale, (int, float)) else None,
        })
    return products

def extract_order_from_sheet(ws, sheet_name):
    """Extract order items, discount, payment terms from an order sheet."""
    date = parse_date(sheet_name)
    items = []
    discount = None
    payment_terms = None
    payment_dates = []
    notes = []

    for r in range(1, ws.max_row + 1):
        code = cv(ws, r, 1)
        desc = cv(ws, r, 2)
        cost = cv(ws, r, 8)
        sale = cv(ws, r, 9)
        qty = cv(ws, r, 10)
        total = cv(ws, r, 11)

        # Look for payment/discount info in col 2 or 9
        for c in range(1, ws.max_column + 1):
            val = cv(ws, r, c)
            if val and isinstance(val, str):
                vl = val.lower().strip()
                # Payment terms like "30-45-60", "CHEGUE 30-45-60", "À VISTA"
                if re.search(r'chegue\s*\d', vl) or re.search(r'^\d+[-/]\d+', vl) or 'à vista' in vl or 'avista' in vl:
                    payment_terms = val.strip()
                # Discount
                if 'desconto' in vl or 'desc.' in vl:
                    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%', val)
                    if m:
                        discount = float(m.group(1).replace(',', '.'))

        # Order items: must have code and qty
        if code is not None and qty is not None:
            if isinstance(code, str) and any(x in code.lower() for x in ['código', 'total', 'chegue', 'prazo']):
                continue
            if isinstance(qty, (int, float)) and qty > 0:
                items.append({
                    "code": str(code).strip().lstrip("'") if code else None,
                    "description": str(desc).strip().lstrip("'") if desc else None,
                    "unit_price": float(sale) if sale and isinstance(sale, (int, float)) else None,
                    "cost_price": float(cost) if cost and isinstance(cost, (int, float)) else None,
                    "qty": int(qty) if isinstance(qty, (int, float)) else qty,
                    "total": float(total) if total and isinstance(total, (int, float)) else None,
                })

    # Also scan for payment dates in columns 10+
    for r in range(1, ws.max_row + 1):
        for c in range(9, ws.max_column + 1):
            val = cv(ws, r, c)
            if val and isinstance(val, str):
                vl = val.strip().lstrip("'")
                # Date patterns like '20.03.26'
                if re.match(r'^\d{2}\.\d{2}\.\d{2,4}$', vl):
                    payment_dates.append(vl)

    return {
        "sheet_name": sheet_name,
        "date": date,
        "discount_pct": discount,
        "payment_terms": payment_terms,
        "payment_dates": payment_dates,
        "items": items,
    }

def process_file(filepath):
    filename = os.path.basename(filepath)
    # Extract client name from filename (remove "PEDIDO AGROPECUARIA " prefix)
    client_name_from_file = re.sub(r'^PEDIDO\s+', '', filename, flags=re.IGNORECASE)
    client_name_from_file = re.sub(r'\.xlsx$', '', client_name_from_file, flags=re.IGNORECASE).strip()

    print(f"\n{'='*80}")
    print(f"FILE: {filename}")
    print(f"CLIENT (from filename): {client_name_from_file}")
    print(f"{'='*80}")

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        return

    print(f"SHEETS: {wb.sheetnames}")

    client_info = None
    products = []
    orders = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sname = sheet_name.strip()

        if sname.upper() in ['PEDIDO MODELO', 'MODELO', 'TABELA DE PRECOS', 'TABELA']:
            print(f"\n--- SHEET: '{sname}' [PRODUCT MODEL] ---")
            client_info = extract_client_from_sheet(ws)
            products = extract_products_from_sheet(ws)
            print(f"  CLIENT INFO: {client_info}")
            print(f"  PRODUCTS ({len(products)} items):")
            for p in products:
                print(f"    {p}")

        elif is_order_sheet(sname):
            print(f"\n--- SHEET: '{sname}' [ORDER] ---")
            order = extract_order_from_sheet(ws, sname)
            orders.append(order)
            print(f"  DATE: {order['date']}")
            print(f"  DISCOUNT: {order['discount_pct']}%")
            print(f"  PAYMENT TERMS: {order['payment_terms']}")
            print(f"  PAYMENT DATES: {order['payment_dates']}")
            print(f"  ITEMS ({len(order['items'])}):")
            for item in order['items']:
                print(f"    {item}")

        elif sname.lower() not in ['plan2', 'plan3', 'sheet1', 'sheet2', 'sheet3', 'planilha1', 'planilha2']:
            print(f"\n--- SHEET: '{sname}' [OTHER] ---")
            # Print first 15 rows
            for r in range(1, 20):
                row_vals = []
                for c in range(1, 20):
                    v = cv(ws, r, c)
                    if v is not None:
                        row_vals.append((c, repr(v)))
                if row_vals:
                    print(f"  Row {r}: {row_vals}")

    # Use filename as client name if not found in sheet
    if client_info and not client_info.get('name'):
        client_info['name'] = client_name_from_file
    elif not client_info:
        client_info = {'name': client_name_from_file}

    return {
        "filename": filename,
        "client": client_info,
        "products": products,
        "orders": orders,
    }

# Collect unique files (avoid duplicates across subdirs)
all_files = {}
for d in DIRS:
    for f in glob.glob(os.path.join(d, "*.xlsx")):
        bname = os.path.basename(f)
        if bname not in all_files:
            all_files[bname] = f

print(f"Found {len(all_files)} unique xlsx files to process")
print("Files:", sorted(all_files.keys()))

results = []
for bname in sorted(all_files.keys()):
    result = process_file(all_files[bname])
    if result:
        results.append(result)

print(f"\n\n{'='*80}")
print(f"SUMMARY: Processed {len(results)} files")
print(f"{'='*80}")
for r in results:
    print(f"\nCLIENT: {r['client']}")
    print(f"  Products: {len(r['products'])}")
    print(f"  Orders: {len(r['orders'])}")
    for o in r['orders']:
        print(f"    Order {o['sheet_name']} ({o['date']}): {len(o['items'])} items, payment={o['payment_terms']}, discount={o['discount_pct']}%")
